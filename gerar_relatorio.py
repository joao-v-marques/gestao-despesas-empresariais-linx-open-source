from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def gerar_relatorio(dados, colunas, destino):
    wb = Workbook()
    ws = wb.active
    ws.title = "relatorio_solicitacoes"

    # Cabeçalho
    header_font = Font(bold=True)

    for i, coluna in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=i, value=coluna)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Linhas de dados
    for linha_idx, linha in enumerate(dados, start=2):
        for col_idx, valor in enumerate(linha, start=1):
            ws.cell(row=linha_idx, column=col_idx, value=valor)

    wb.save(destino)